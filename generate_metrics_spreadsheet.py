#!/usr/bin/python

import sys
import os
import csv
import sqlite3
import operator
import calendar
from datetime import date
from datetime import datetime
import openpyxl
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.styles import Font, Color, Alignment, colors
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

full_month_dict = {
        1 : 'January',
        2 : 'February',
        3 : 'March',
        4 : 'April',
        5 : 'May',
        6 : 'June',
        7 : 'July',
        8 : 'August',
        9 : 'September', 
        10 : 'October',
        11 : 'November',
        12 : 'December'
    }

translate_severity_dict = {
        "S0" : "S0",
        "S1" : "S1",
        "S2" : "S2",
        "S3" : "S3",
        "L1" : "S0",
        "L2" : "S1",
        "L3" : "S2",
        "L4" : "S3",
        "Unprioritized" : "S3",
        "Low" : "S3",
        "None" : "S3",
        "Trivial" : "S3",
        "Minor" : "S2",
        "Medium" : "S2",
        "Normal" : "S2",
        "High" : "S1",
        "Major" : "S1",
        "Urgent" : "S0",
        "Critical" : "S0",
        "Blocker" : "S0"
}
    
def generate_products_dict():
    # return product dic
    # product_code : [product_name, business_unit, dev_owner, qe_owner]
    try:
        '''
        product_name char(30),
        product_code char(30),
        business_unit char(20),
        dev_owner char(20),
        qe_owner char(20)

        '''
        dic = {}
        with open("product_mapping.csv", "r") as f:
            lines = f.readlines()
            counter = 1
            for line in lines:
                arr = line.strip('\n').split("|")
                if arr[1] == 'CODE':
                    arr[1] = 'CODE'+str(counter)
                    counter+=1
                dic[arr[1]] = [arr[0], arr[2], arr[3], arr[4]]
                print arr[1], [arr[0], arr[2], arr[3], arr[4]]

        return dic

    except Exception as e:
        print "Got exception in generate_products_dict : " + str(e)
        sys.exit()

products_dict = generate_products_dict()

def flatten_counts_into_array(severity_array, sev_code, count):
    try:
        if sev_code == 'S0':
            severity_array[0] = count
        elif sev_code == 'S1':
            severity_array[1] = count
        elif sev_code == 'S2':
            severity_array[2] = count
        elif sev_code == 'S3':
            severity_array[3] = count
        else:
            print 'WARNING: Received request without severity code assigned.  Assigning S3 '
            #severity_array[3] = severity_array[3] + count
        return severity_array

    except Exception as e:
        print "Got exception in flatten_counts_into_array : " + str(e)
        sys.exit()

# Delete old rows and import new jira exported file
def import_jira_bugs(db_file, jira_csv_filename, products_dict):
    print "Inside import_jira_bugs"
    try:
        conn = sqlite3.connect(db_file)
        print "import_jira_bugs - after connect"
        conn.text_factory = str
        conn.execute("delete from all_issues")
        conn.commit()
        print "import_jira_bugs - after delete commit"

        with open(jira_csv_filename, 'rb') as f:
            reader = csv.reader(f)
            for row in reader:  # throw away header line
                #print "import_jira_bugs - in loop " + str(row[8])
                if row[0] == 'Issue Type': continue  # skip header line
                issue_type = row[0]
                issue_key  = row[1]
                issue_id   = row[2]
                summary    = row[3]
                status     = row[4]
                created_date = datetime.strptime(row[5], '%d/%b/%y %I:%M %p')
                created = created_date.date()
                updated_date = datetime.strptime(row[6], '%d/%b/%y %I:%M %p')
                updated = updated_date.date()
                security_severity = row[7]
                components  = row[14]
                if len(components.strip()) == 0: # to-do Not SE part will get product from other place
                    key = issue_key.split('-')[0]
                    if key and key in products_dict:
                        components = products_dict[key][0]
                components2 = row[15]
                if issue_key.split('-')[0] == 'SE':
                    priority = 'SE'
                else:
                    priority = 'QE'
                resolved_date = None
                if row[17]:
                    resolved_date = datetime.strptime(row[17], '%d/%b/%y %I:%M %p')
                when_discovered = "" 
                get_severity_level = "" 
                to_insert = (issue_type, 
                    issue_key, 
                    issue_id, 
                    summary, 
                    status, 
                    created, 
                    updated, 
                    security_severity, 
                    components,
                    priority,
                    resolved_date,
                    when_discovered,
                    get_severity_level)
                conn.execute('insert into all_issues values(?,?,?,?,?,?,?,?,?,?,?,?,?)', to_insert)

        conn.commit()
        conn.close()
        print "import_jira_bugs - after conn close"
    except Exception as e:
        print "Got exception in import_jira_bugs  " + str(e)
        sys.exit()

def translate_severity(severity_list):
    severity = ""
    for i in severity_list:
        i = i.strip()
        if len(i) != 0:
            severity = translate_severity_dict[i]
    return severity

# Display db count after import
def display_number_of_issues_imported_count(db_file):
    conn = sqlite3.connect(db_file)
    cursor = conn.execute("select count(*) from all_issues;")
    for row in cursor:
        print "display_number_of_issues_imported_count = ", str(row[0])
    conn.close()


# Total Current Security Bug Count as of November 30, 2016
def get_total_current_security_bug_count_as_of(db_file, end_of_month, Dep):
    severity_array = [0, 0, 0, 0]

    try:
        conn = sqlite3.connect(db_file)
        if Dep == "":
            sql = "select security_severity, count(*) from all_issues where status != 'Closed' and created <= '%s' group by security_severity;" % (end_of_month,)
        else:
            sql = "select security_severity, count(*) from all_issues where priority = '%s' and status != 'Closed' and created <= '%s' group by security_severity;" % (Dep, end_of_month,)
        cursor = conn.execute(sql)
        for row in cursor:
            #print "security_severity = ", row[0]
            #print "count = ", str(row[1])
            severity_array = flatten_counts_into_array(severity_array, row[0], row[1])
        conn.close()
    except Exception as e:
        print "Got exception in get_total_current_security_bug_count_as_of  " + str(e)
        sys.exit()

    print "  in get total  severity_array=" + str(severity_array)
    return severity_array



# Total Security Bugs Opened in November,  2016  - date format : '2016-11-01'
def get_total_security_bugs_opened_in_month(db_file, start_of_month, end_of_month):
    severity_array = [0, 0, 0, 0]

    try:
        conn = sqlite3.connect(db_file)
        sql = "select security_severity, count(*) from all_issues where created >= '%s' and created <= '%s' group by security_severity;" % (start_of_month, end_of_month)
        cursor = conn.execute(sql)
        for row in cursor:
            #print "security_severity = ", row[0]
            #print "count = ", row[1]
            severity_array = flatten_counts_into_array(severity_array, row[0], row[1])
        conn.close()
    except Exception as e:
        print "Got exception in get_total_security_bugs_opened_in_month  " + str(e)
        sys.exit()

    return severity_array


# Total Security Bugs Closed in November 2016
def get_total_security_bugs_closed_in_month(db_file, start_of_month, end_of_month):
    severity_array = [0, 0, 0, 0]

    try:
        conn = sqlite3.connect(db_file)
        sql = "select security_severity, count(*) from all_issues where status = 'Closed' and updated >= '%s' and updated <= '%s' group by security_severity;" % (start_of_month, end_of_month)
        cursor = conn.execute(sql)
        for row in cursor:
            #print "security_severity = ", row[0]
            #print "count = ", row[1] severity_array = flatten_counts_into_array(severity_array, row[0], row[1])
            severity_array = flatten_counts_into_array(severity_array, row[0], row[1])
        conn.close()
    except Exception as e:
        print "Got exception in get_total_security_bugs_closed_in_month: " + str(e)
        sys.exit()

    return severity_array


def get_top10_by_bug_severity(db_file, end_of_month):
    
    # dict of array - key of dict is the Application.  The array contains the count for S0, S1, etc
    issue_count_by_app = {}
    prev_app = ''
    severity_array = [0, 0, 0, 0]
    
    try:
        conn = sqlite3.connect(db_file)
        sql = "select components, security_severity, count(*) from all_issues where status != 'Closed' and created <= '%s' group by components, security_severity;" % (end_of_month,)
        #sql = "select components, security_severity, count(*) from all_issues where priority = 'SE' and status = 'Open' and created <= '%s' group by components, security_severity;" % (end_of_month,)
        cursor = conn.execute(sql)
        for row in cursor:
            #print "component = ", row[0]
            #print "security_severity = ", row[1]
            #print "count = ", row[2], "\n"
            app = row[0]
            if not app or len(app) == 0:
                continue
            #print app
        
            if app == prev_app:
                if row[1] == 'S0':
                    severity_array[0] = row[2]
                elif row[1] == 'S1':
                    severity_array[1] = row[2]
                elif row[1] == 'S2':
                    severity_array[2] = row[2]
                elif row[1] == 'S3':
                    severity_array[3] = row[2]
                else:
                    print 'found row without severity assigned in app=' + app
            else:
                if prev_app != '':  # break and not first time
                    issue_count_by_app[prev_app] = severity_array
        
                severity_array = [0, 0, 0, 0]
                if row[1] == 'S0':
                    severity_array[0] = row[2]
                elif row[1] == 'S1':
                    severity_array[1] = row[2]
                elif row[1] == 'S2':
                    severity_array[2] = row[2]
                elif row[1] == 'S3':
                    severity_array[3] = row[2]
                else:
                    print 'found row without severity assigned in app=' + app
        
                prev_app = app
         
        conn.close()
    except Exception as e:
        print "Got exception in get_top10_by_bug_severity  " + str(e)
        sys.exit()


    # store last app counts
    issue_count_by_app[prev_app] = severity_array
    
    
    
    # the following dict keys on appname and stores the weight as the value
    top10_app_by_severity = {}
    for app in issue_count_by_app:
        severity_array = issue_count_by_app[app]
        weight = severity_array[0] * 1000000000 + severity_array[1] * 1000000 + severity_array[2] * 1000 + severity_array[3]
        top10_app_by_severity[app] = weight
    
    
    # now print the results in decreasing weight order
    top10 = []
    idx = 0
    sorted_x = sorted(top10_app_by_severity.items(), key=operator.itemgetter(1), reverse=True)
    for (app, weight) in sorted_x:
        bug_count_array = issue_count_by_app[app]
        #print app + '   S0=' + str(bug_count_array[0])  + ' S1=' + str(bug_count_array[1]) + ' S2=' + str(bug_count_array[2]) + ' S3=' + str(bug_count_array[3]) + '  weight=' + str(weight)
        #top10[app] = bug_count_array
        top10.append( [app] + bug_count_array )
        idx = idx + 1
        if idx > 9:
            break

    return top10  # list of lists.  the latter has the app as the first element and the counts following it



def get_open_issues_list(db_file, end_of_month):
    all_rows = []

    try:
        conn = sqlite3.connect(db_file)
        sql = "select summary, created, security_severity, components, 'https://jira.rax.io/browse/' || issue_key from all_issues where status !='Closed' and created <= '%s' order by security_severity;" % (end_of_month,)
        cursor = conn.execute(sql)
        for row in cursor:
            all_rows.append(row)
        
        conn.close()
    except Exception as e:
        print "Got exception in get_open_issues_list  " + str(e)
        sys.exit()

    return all_rows

def get_numeric_month(month):
    month_dict = {
        'jan' : 1,
        'feb' : 2,
        'mar' : 3,
        'apr' : 4,
        'may' : 5,
        'jun' : 6,
        'jul' : 7,
        'aug' : 8,
        'sep' : 9, 
        'oct' : 10,
        'nov' : 11,
        'dec' : 12
    }
    return month_dict[month]

def get_history_array(history_file):
    # read file from the end
    last_12_months_totals = []
    f = open(history_file)
    lines = f.readlines()
    size = len(lines)
    if size > 12:
        lines = lines[size-12:]
    for line in reversed(lines):
        line = line.strip()
        history_array = line.split(",")
        last_12_months_totals.append(history_array)
    f.close()
    return last_12_months_totals

def write_history_file(history_file, report_date, opened_in_month_array, total_opened_in_month, closed_in_month, total_current_security_bug_count_as_of, open_to_date):
    print "write_history_file", history_file,'1', report_date,'2',opened_in_month_array,'3', total_opened_in_month,'4', closed_in_month,'5', total_current_security_bug_count_as_of,'6', open_to_date
    with open(history_file, 'r+') as f:
        lines = f.readlines()
        if lines[-1].split(',')[0].strip() == report_date.strip():
            print "find duplicate"
            f.seek(0, os.SEEK_END)
            pos = f.tell() - 1
            while pos > 0 and f.read(1) != "\n":
                pos -= 1
                f.seek(pos, os.SEEK_SET)
            if pos > 0:
                f.seek(pos, os.SEEK_SET)
                f.truncate()
            f.write('\n')

        print "write to file"
        data = report_date + ',' + \
            str(opened_in_month_array).strip(' []') + ',' + \
            str(total_opened_in_month) + ',' + \
            str(closed_in_month) + ',' + \
            str(total_current_security_bug_count_as_of).strip(' []') + ',' + \
            str(open_to_date) + '\n'
        print "write to file: ", data
        f.write(data)


def retreve_data_save_to_db(year, month, username, db_file):

    print "======== generating report for month=" + month + ' ' + year
    numeric_month = get_numeric_month(month)
    last_day_of_month = calendar.monthrange(int(year), numeric_month)[1]
    
    password = os.environ['PASSWORD']
    export_file_name = 'IPA_metrics_' + str(numeric_month) + '_' + str(last_day_of_month) + '_' + year + '.csv'
    print 'export_file_name=' + export_file_name
    #os.system('wget -O ' + export_file_name + ' --ignore-length=on "https://jira.rax.io/sr/jira.issueviews:searchrequest-csv-current-fields/19350/SearchRequest-19350.csv?tempMax=2000&os_username=' + username + '&os_password=' + password + '"')

    
    #generage products_dict and pass to import_jira_bugs

    # import csv into sqlite
    import_jira_bugs(db_file, export_file_name, products_dict)
    
    # for debugging
    #display_number_of_issues_imported_count(db_file)
    
#    new_spreadsheet_name = 'IPA_metrics_' + str(numeric_month) + '_' + str(last_day_of_month) + '_' + year + '.xlsx'
#print 'new_spreadsheet_name=' + new_spreadsheet_name
#os.system('cp ./IPA_metrics_TEMPLATE.xlsx ./' + new_spreadsheet_name)
#os.system('cp ./exceptions_IPA_metrics_TEMPLATE.xlsx ./exceptions_' + new_spreadsheet_name)


#----------------------------------------------------
# Create main spreadsheet
#----------------------------------------------------
#wb = xw.Workbook(spreadsheet, app_visible=None)
#wb = xw.Book('./' + new_spreadsheet_name)
#wb.set_current()
#metrics_sheet = wb.sheets['Metrics']

#----------------------------------------------------
# Create exceptions spreadsheet
#----------------------------------------------------
#exceptions_wb = xw.Book('./' + 'exceptions_' + new_spreadsheet_name)
#exceptions_sheet = exceptions_wb.sheets['Exceptions']

#----------------------------------------------------
# Create total current security open bug count as of
#----------------------------------------------------

def get_end_of_month(month):
    numeric_month = get_numeric_month(month)
    last_day_of_month = calendar.monthrange(int(year), numeric_month)[1]
    
    if numeric_month > 9:  # prefix with 0 if 1 digit month
        end_of_month = year + '-' + str(numeric_month) + '-' + str(last_day_of_month)
    else:
        end_of_month = year + '-' + '0' + str(numeric_month) + '-' + str(last_day_of_month)

    return end_of_month

def get_start_of_month(month):
    numeric_month = get_numeric_month(month)
    if numeric_month > 9:  # prefix with 0 if 1 digit month
        start_of_month = year + '-' + str(numeric_month) + '-01'
    else:
        start_of_month = year + '-' + '0' + str(numeric_month) + '-01'
    return start_of_month


def create_total_current_open_bug_count(year, month, bigArr, bigTitle, db_file, Dep):
    numeric_month = get_numeric_month(month)
    last_day_of_month = calendar.monthrange(int(year), numeric_month)[1]
    
    if numeric_month > 9:  # prefix with 0 if 1 digit month
        end_of_month = year + '-' + str(numeric_month) + '-' + str(last_day_of_month)
    else:
        end_of_month = year + '-' + '0' + str(numeric_month) + '-' + str(last_day_of_month)
    
    severity_array = get_total_current_security_bug_count_as_of(db_file, end_of_month, Dep)
    sevarr = severity_array[:]
    sevarr.append(sum(severity_array))
    
    s_sev = ['S0', 'S1', 'S2', 'S3', 'Total']
    bigArr.append([s_sev, sevarr])
    if Dep != "":
        Dep = " " + Dep
    title = 'Total Current' + Dep + ' Bug Count as of ' + full_month_dict[numeric_month] + ' ' + str(last_day_of_month) + ', ' + year
    bigTitle.append(title)
    
    open_to_date = 0
    for count in severity_array:
        open_to_date += count
    #for idx,count in enumerate(severity_array):
    #    print '  count=' + str(count)

    return open_to_date, severity_array
    

#----------------------------------------------------
# Create total security bugs opened in month
#----------------------------------------------------
def create_total_bugs_opened_in_month(year, month, bigArr, bigTitle, db_file):
    numeric_month = get_numeric_month(month)
#    if numeric_month > 9:  # prefix with 0 if 1 digit month
#        start_of_month = year + '-' + str(numeric_month) + '-01'
#    else:
#        start_of_month = year + '-' + '0' + str(numeric_month) + '-01'

    start_of_month = get_start_of_month(month)
    end_of_month = get_end_of_month(month)
    
    severity_array = get_total_security_bugs_opened_in_month(db_file, start_of_month, end_of_month)
    
    sevarr = severity_array[:]
    sevarr.append(sum(severity_array))
    s_sev = ['S0', 'S1', 'S2', 'S3', 'Total']
    bigArr.append([s_sev, sevarr])
    title = 'Total Bugs Opened in ' + full_month_dict[numeric_month] + ' ' + year
    bigTitle.append(title)
    
    opened_in_month_array = severity_array[:]
    return opened_in_month_array

#----------------------------------------------------
# Create total security bugs closed in month
#----------------------------------------------------

def create_total_bugs_closed_in_month(year, month, bigArr, bigTitle, db_file):

    numeric_month = get_numeric_month(month)
    title = 'Total Bugs Closed in ' + full_month_dict[numeric_month] + ' ' + year

    start_of_month = get_start_of_month(month)
    end_of_month = get_end_of_month(month)

    severity_array =  get_total_security_bugs_closed_in_month(db_file, start_of_month, end_of_month)
    
    sevarr = severity_array
    sevarr.append(sum(severity_array))
    s_sev = ['S0', 'S1', 'S2', 'S3', 'Total']
    bigArr.append([s_sev, sevarr])
    bigTitle.append(title)

    closed_in_month = 0
    for count in severity_array:
        closed_in_month += count

    return closed_in_month

#----------------------------------------------------
# Create top 10 by bug severity section
#----------------------------------------------------
#product_column = 'G'
#s0_column = 'H'
#idx = 3

def create_top_10_by_bug_severity_section(year, month, bigArr, bigTitle, db_file):
    title = "Top 10 by Bug Severity"
    arr = []
    end_of_month = get_end_of_month(month)
    # top10 is a list of lists.  the latter has the app as the first element and the counts following it
    top10 = get_top10_by_bug_severity(db_file, end_of_month)
    for elem in top10:
        app = elem.pop(0)
        bug_count_array = elem
    
        t = elem
        t.append(sum(elem))
        t = [app] + t
    
        arr.append(t)
    
    s_sev = ['Product', 'S0', 'S1', 'S2', 'S3', 'Total']
    arr.insert(0, s_sev)
    bigArr.append(arr)
    bigTitle.append(title)
    
    
#----------------------------------------------------
def create_trending_section(bigArr, bigTitle, history_file):
    
    title = "Trending Total Bug Count By Month"
    se_v = ["Month", "S0", "S1", "S2", "S3", "Opened in Month", "Closed in Month"]
    app = []
    app.append(se_v)
    
    # Create trending section
    #idx = 3
    last_12_months_totals = get_history_array(history_file)
    
    for history_array in reversed(last_12_months_totals):
        print history_array[:7]
        app.append(history_array[:7])
    
    bigArr.append(app)
    bigTitle.append(title)

def create_cumulative_trending_section(bigArr, bigTitle):
    history_file = 'trending_totals_by_month.csv'
    last_12_months_totals = get_history_array(history_file)
    title = "Trending Total Bug Count"
    se_v = ["Month", "S0", "S1", "S2", "S3", "Total Open as of Month"]
    arr = []
    arr.append(se_v)
    for history_array in reversed(last_12_months_totals):
        t = []
        t.append(history_array[0])
        t.append(history_array[7])
        t.append(history_array[8])
        t.append(history_array[9])
        t.append(history_array[10])
        t.append(history_array[11])
        arr.append(t)
    bigArr.append(arr)
    bigTitle.append(title)
    
#----------------------------------------------------
# Create open issues sheet
#----------------------------------------------------
def create_open_issues_sheet(year, month, db_file):
    end_of_month = get_end_of_month(month)
    all_rows = get_open_issues_list(db_file, end_of_month)

    print "all_rows : ", len(all_rows)

    arr = []
    for row in all_rows:
        created_date_str = row[1]
        created_date = datetime.strptime(created_date_str, '%Y-%m-%d')
        d0 = datetime.strptime(end_of_month , '%Y-%m-%d').date()
        d1 = created_date.date()
        age = d0 - d1

        product_code = row[4].split("/")[-1].split("-")[0]
        if product_code in products_dict:
            product_name, business_unit, dev_owner, qe_owner = products_dict[product_code]
        else:
            product_name, business_unit, dev_owner, qe_owner = None, None, None, None
    
        arr.append([row[0], created_date_str, row[2], age.days, row[3], row[4], business_unit, dev_owner, qe_owner])
    return arr


    
def test_my_way(ws, bigArr, bigTitle):
    try:
        row = 1
        col = 1

        chardata = bigArr[3][:]
        chardata.append(bigArr[4][1])

        char = [[] for i in range(6)]
        char[0] = ["Severity", "SE", "QE"]
        for i in range(len(chardata)):
            for j in range(len(chardata[0])):
                char[j+1].append(chardata[i][j])


        write_arr_in_sheet(ws, 18, 7, "", char)

        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "Severity Chart"
        chart1.y_axis.title = 'Bug Number'
        chart1.x_axis.title = 'Severity'
        
        cats = Reference(ws, min_col=7, min_row=19, max_row=23)
        data = Reference(ws, min_col=8, max_col=9, min_row=18, max_row=23)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.shape = 6
        ws.add_chart(chart1, "G18")

        
        for i in range(5):
            write_arr_in_sheet(ws, row, col, bigTitle[i], bigArr[i])
            row+=4
        
        row = 1
        col += len(bigArr[0][1])+1
        write_arr_in_sheet(ws, row, col, bigTitle[5], bigArr[5])

        row = 1
        col += len(bigArr[5][0])+1
        t = col
        write_arr_in_sheet(ws, row, col, bigTitle[6], bigArr[6])

        row = 15
        col = t
        write_arr_in_sheet(ws, row, col, bigTitle[7], bigArr[7])
        
    except:
        print "Unexpected error:", sys.exc_info()
        print "something wrong"

def second_sheet(ws, arr):
    try:
        title = ""
        first_line = ["Name", "Date", "Severity", "Age", "Product", "Link", "Business Unit", "Dev Owner", "QE Owner"]
        arr.insert(0, first_line)
        write_arr_in_sheet(ws, 1, 1, title, arr, True)

    except:
        print "Unexpected error:", sys.exc_info()

def write_arr_in_sheet(ws, row, col, title, arr, link=False):
    try:    
        start_col = chr(ord('A')+col-1)
        end_col = chr(ord('A')+col+len(arr[0])-2)
        start_row = row
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        redFill = PatternFill(start_color='7e0207', end_color='7e0207', fill_type='solid')
        
        if title != "":
            ws.merge_cells(start_col+str(row)+':'+end_col+str(row))
            ws.column_dimensions[chr(ord('A')+col-1)].width = len(title)+2
            ws.cell(row=start_row, column=col).value = title 
            ws.cell(row=start_row, column=col).alignment = Alignment(horizontal='center')
            ws.cell(row=start_row, column=col).font = Font(bold=True)
            start_row += 1

        ma = [10] * len(arr[1])
        for i in range(len(arr)):
            for j in range(len(arr[0])):
                if type(arr[i][j]) == unicode:
                    arr[i][j] = arr[i][j].encode('ascii','ignore')
                ma[j] = max(ma[j], len(str(arr[i][j])))

        for i in range(len(arr[0])):
            ws.column_dimensions[chr(ord('A')+i+col-1)].width = ma[i] 

        for i in range(len(arr[0])):
            ws.cell(row=start_row, column=col+i).value = arr[0][i] 
            ws.cell(row=start_row, column=col+i).font = Font(bold=True, color=colors.WHITE)
            ws.cell(row=start_row, column=col+i).alignment = Alignment(horizontal='center')
            ws.cell(row=start_row, column=col+i).fill = redFill 
            ws.cell(row=start_row, column=col+i).border = thin_border 
        
        start_row += 1
        for i in range(len(arr)-1):
            for j in range(len(arr[1])):
                if j != 0:
                    ws.cell(row=start_row+i, column=col+j).alignment = Alignment(horizontal='right')
                if j == len(arr[1])-1 and link:
                    ws.cell(row=start_row+i, column=col+j).value = '=HYPERLINK("{}", "{}")'.format(str(arr[i+1][j]), arr[i+1][j]) 
                else:
                    ws.cell(row=start_row+i, column=col+j).value = arr[i+1][j]

    except Exception as e:
        print "Unexpected error:", sys.exc_info(), str(e)
        print "write_arr_in_sheet something wrong"
# --------------------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------------------

if __name__ == "__main__":

    if len(sys.argv) < 4:
        print 'Usage:  python generate_metrics_spreadsheet.py <month> <year> <jira_username> <db_file>'
        quit()

    month = sys.argv[1]
    year = sys.argv[2]
    username = sys.argv[3]
    db_file = sys.argv[4]

    if not month or month not in ('jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec'):
        print 'Invalid month supplied.  Supply one of: jan, feb, mar, apr, may, jun, jul, aug, sep, oct, nov, dec'
        quit()

    if year and (int(year) == date.today().year or int(year) == date.today().year - 1):
        pass
    else:
        print 'Invalid year supplied.  It needs to be this year or last year.'
        quit()

    if not username:
        print 'Invalid username supplied.'
        quit()

    if not db_file:
        print 'Invalid db_file supplied.'
        quit()

    history_file = 'trending_totals_by_month.csv'
    report_date = month.capitalize() + '-' + year

    filename = 'Report.xlsx'
    wb = openpyxl.Workbook()

    remove_sheet_name = wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(remove_sheet_name)

    retreve_data_save_to_db(year, month, username, db_file)

    Departments = ["SE", "QE"]

    bigArr = []
    bigTitle = []

    open_to_date, total_current = create_total_current_open_bug_count(year, month, bigArr, bigTitle, db_file, "")

    opened_in_month_array = create_total_bugs_opened_in_month(year, month, bigArr, bigTitle, db_file)

    closed_in_month = create_total_bugs_closed_in_month(year, month, bigArr, bigTitle, db_file)

    for Dep in Departments:
        create_total_current_open_bug_count(year, month, bigArr, bigTitle, db_file, Dep)

    create_top_10_by_bug_severity_section(year, month, bigArr, bigTitle, db_file)
    
    total_opened_in_month = sum(opened_in_month_array)

    write_history_file(history_file, report_date, opened_in_month_array, total_opened_in_month, closed_in_month, total_current, open_to_date)

    create_trending_section(bigArr, bigTitle, history_file)

    arr = create_open_issues_sheet(year, month, db_file)

    create_cumulative_trending_section(bigArr, bigTitle)

    ws = wb.create_sheet("Sheet")
    test_my_way(ws, bigArr, bigTitle)
    
    sheet_name = "Open Findings"
    ws = wb.create_sheet(sheet_name)
    second_sheet(ws, arr)

    wb.save(filename)
